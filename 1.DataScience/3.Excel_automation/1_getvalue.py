import openpyxl

path = '/Users/bactran/Documents/1.Learning/2.Python/1.DataScience/3.Excel_automation/'

file_names = ['SampleData.xlsx','SampleData2.xlsx']

excel_files = [path + i for i in file_names]

values = []

# get values in a cell
for file in excel_files: 
    # open workbook
    workbook = openpyxl.load_workbook(file)

    # select worksheet
    worksheet = workbook['SalesOrders']

    # get cell value
    cell_value = worksheet['G11'].value
    values.append(cell_value)

    print(cell_value)

# Create new sheets
for file in excel_files:
    workbook = openpyxl.load_workbook(file)
    if 'NewSheet' not in workbook.sheetnames:
        workbook.create_sheet('NewSheet')
        workbook.save(file)
